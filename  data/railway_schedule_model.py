import pandas as pd
from ortools.sat.python import cp_model

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


def calc_excel(data_file):
    """
    读取Excel文件并计算三个指标：
    1. 过修程度指标（Z检修和L检修）
    2. 换车次数指标
    3. 检修均衡指标（Z检修和L检修）
    """
    # 读取数据文件
    df_schedule = pd.read_excel(data_file, sheet_name="甘特图")

    # 读取车组初始信息
    df_mileage = pd.read_excel("Data.xlsx", sheet_name="车组里程修时信息")
    df_recover = pd.read_excel("Data.xlsx", sheet_name="车组修后恢复信息")

    # 获取恢复值
    recover_Z_day = int(df_recover[df_recover['检修类型'] == 'Z']['修后恢复天数'].iloc[0])
    recover_Z_km = int(df_recover[df_recover['检修类型'] == 'Z']['修后恢复公里数'].iloc[0])
    recover_L_km = int(df_recover[df_recover['检修类型'] == 'L']['修后恢复公里数'].iloc[0])

    # 获取车组初始剩余数据
    initial_Z_day = dict(zip(df_mileage['车组号'], df_mileage['Z剩余天数']))
    initial_Z_km = dict(zip(df_mileage['车组号'], df_mileage['Z剩余里程']))
    initial_L_km = dict(zip(df_mileage['车组号'], df_mileage['L剩余里程']))

    # 1. 计算过修程度指标
    z_over_repair_day = 0
    z_over_repair_km = 0
    l_over_repair_km = 0

    # 获取所有执行Z检修的车组
    z_maint_days = [col for col in df_schedule.columns if col.lower().startswith('day')]
    z_maint_vehicles = []

    for day in z_maint_days:
        z_cells = df_schedule[df_schedule['任务'] == 'Z'][day].values[0]
        if pd.notna(z_cells) and z_cells != '':
            for v in z_cells.split(','):
                z_maint_vehicles.append((v, day))

    # 计算Z检修过修指标
    for v, day in z_maint_vehicles:
        # 获取修前剩余天数
        z_over_repair_day += initial_Z_day[v]
        # 获取修前剩余里程
        z_over_repair_km += initial_Z_km[v]

    z_over_repair_indicator = (
                (z_over_repair_day / recover_Z_day + z_over_repair_km / recover_Z_km) / 2) if z_maint_vehicles else 0

    # 获取所有执行L检修的车组
    l_maint_vehicles = []
    for day in z_maint_days:
        l_cells = df_schedule[df_schedule['任务'] == 'L'][day].values[0]
        if pd.notna(l_cells) and l_cells != '':
            for v in l_cells.split(','):
                l_maint_vehicles.append((v, day))

    # 计算L检修过修指标
    for v, day in l_maint_vehicles:
        l_over_repair_km += initial_L_km[v]

    l_over_repair_indicator = l_over_repair_km / recover_L_km if l_maint_vehicles else 0



    # 2. 计算换车次数指标
    # 读取交路信息
    df_route = pd.read_excel("Data.xlsx", sheet_name="待排交路信息")
    route_rid = dict(zip(df_route['交路'], df_route['R_ID']))

    # 获取R_ID对应的交路列表
    from collections import defaultdict
    r_id_routes = defaultdict(list)
    for r, rid in route_rid.items():
        r_id_routes[rid].append(r)

    total_change = 0
    max_possible_change = len(z_maint_days) - 1  # day-1

    for rid, r_list in r_id_routes.items():
        # if len(r_list) < 2:  # 单交路不需要连续执行
        #     continue

        last_route = r_list[-1]  # 当前交路组的最后一个交路
        first_route = r_list[0]  # 当前交路组的第一个交路

        change_count = 0
        for t in range(1, len(z_maint_days)):
            prev_day = z_maint_days[t - 1]
            curr_day = z_maint_days[t]

            # 获取前一天执行最后交路的车组
            prev_vehicle = None
            for v in df_mileage['车组号']:
                if df_schedule.loc[df_schedule['任务'] == last_route, prev_day].values[0] == v:
                    prev_vehicle = v
                    break

            # 获取当天执行首交路的车组
            curr_vehicle = None
            for v in df_mileage['车组号']:
                if df_schedule.loc[df_schedule['任务'] == first_route, curr_day].values[0] == v:
                    curr_vehicle = v
                    break

            if prev_vehicle and curr_vehicle and prev_vehicle != curr_vehicle:
                change_count += 1

        if max_possible_change > 0:
            total_change += change_count / max_possible_change

    change_indicator = total_change

    # 3. 计算检修均衡指标
    # 计算每日Z检修量
    z_daily_counts = []
    for day in z_maint_days:
        z_cells = df_schedule[df_schedule['任务'] == 'Z'][day].values[0]
        count = 0 if pd.isna(z_cells) or z_cells == '' else len(z_cells.split(','))
        z_daily_counts.append(count)

    # 计算Z检修方差
    z_mean = sum(z_daily_counts) / len(z_daily_counts) if z_daily_counts else 0
    z_variance = sum((x - z_mean) ** 2 for x in z_daily_counts) / len(z_daily_counts) if z_daily_counts else 0

    # 计算每日L检修量
    l_daily_counts = []
    for day in z_maint_days:
        l_cells = df_schedule[df_schedule['任务'] == 'L'][day].values[0]
        count = 0 if pd.isna(l_cells) or l_cells == '' else len(l_cells.split(','))
        l_daily_counts.append(count)

    # 计算L检修方差
    l_mean = sum(l_daily_counts) / len(l_daily_counts) if l_daily_counts else 0
    l_variance = sum((x - l_mean) ** 2 for x in l_daily_counts) / len(l_daily_counts) if l_daily_counts else 0

    # 打印结果
    print("计算结果:")
    print(f"1. 过修程度指标:")
    print(f"   Z检修过修指标: {z_over_repair_indicator:.4f}")
    print(f"   L检修过修指标: {l_over_repair_indicator:.4f}")
    print(f"2. 换车次数指标: {change_indicator:.4f}")
    print(f"3. 检修均衡指标:")
    print(f"   Z检修均衡指标: {z_variance:.4f}")
    print(f"   L检修均衡指标: {l_variance:.4f}")

    return {
        "z_over_repair": z_over_repair_indicator,
        "l_over_repair": l_over_repair_indicator,
        "change_indicator": change_indicator,
        "z_variance": z_variance,
        "l_variance": l_variance
    }


def export_to_excel(solver, x, z, l, vehicles, routes, days, filename="排班结果.xlsx"):
    """
    将排班结果导出到Excel文件

    参数:
        solver: 求解器对象
        x: 交路分配变量字典
        z: Z检修变量字典
        l: L检修变量字典
        vehicles: 车组列表
        routes: 交路列表
        days: 日期列表
        filename: 输出文件名
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "甘特图"

    # 写入表头
    headers = ["任务"] + days
    ws.append(headers)

    # 设置表头样式
    header_font = Font(bold=True)
    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = header_font

    # 1. 写入交路排班
    for r in routes:
        row_data = [r]
        for d in days:
            assigned = ""
            for v in vehicles:
                if solver.Value(x[v, r, d]):
                    assigned = v
                    break
            row_data.append(assigned)
        ws.append(row_data)

    # 2. 写入Z检修
    z_row = ["Z"]
    for d in days:
        z_list = [v for v in vehicles if solver.Value(z[v, d])]
        z_row.append(",".join(z_list))
    ws.append(z_row)

    # 3. 写入L检修
    l_row = ["L"]
    for d in days:
        l_list = [v for v in vehicles if solver.Value(l[v, d])]
        l_row.append(",".join(l_list))
    ws.append(l_row)

    # 设置单元格对齐和列宽
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    for col in range(1, len(days) + 2):
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].width = 15

    # 保存文件
    wb.save(filename)
    print(f"排班结果已保存到 {filename}")


def main():
    # 读取数据
    data_file = "Data.xlsx"
    df_mileage = pd.read_excel(data_file, sheet_name="车组里程修时信息")
    df_route = pd.read_excel(data_file, sheet_name="待排交路信息")
    df_day1 = pd.read_excel(data_file, sheet_name="Day1检修上线情况")
    df_capacity = pd.read_excel(data_file, sheet_name="班组检修能力")
    df_candidate = pd.read_excel(data_file, sheet_name="候选交路", header=None)
    df_recover = pd.read_excel(data_file, sheet_name="车组修后恢复信息")

    # 参数定义
    vehicles = list(df_mileage['车组号'])
    routes = list(df_route['交路'])
    route_idx = {r: i for i, r in enumerate(routes)}
    days = [f'day{i}' for i in range(1, 9)]  # day1~day8
    num_days = len(days)

    # 路线是否需要运行（交路调度需求矩阵）
    need_route = {(r, d): int(df_route.loc[route_idx[r], d]) for r in routes for d in days}

    # 获取R_ID映射
    route_rid = dict(zip(df_route['交路'], df_route['R_ID']))

    # 获取连续交路（R_ID对应的交路列表）
    from collections import defaultdict
    r_id_routes = defaultdict(list)
    for r, rid in route_rid.items():
        r_id_routes[rid].append(r)

    # 车组候选交路集合（按R_ID）
    vehicle_candidates = {}
    for i, row in df_candidate.iterrows():
        v = row[0]
        candidate_rids = row.dropna().iloc[1:].tolist()
        candidate_routes = []
        for rid in candidate_rids:
            candidate_routes.extend(r_id_routes.get(rid, []))
        vehicle_candidates[v] = list(set(candidate_routes))

    # 获取检修能力
    capacity_Z = dict(zip(days, df_capacity.loc[df_capacity['maintlevel'] == 'Z'].values[0][1:]))
    capacity_L = dict(zip(days, df_capacity.loc[df_capacity['maintlevel'] == 'L'].values[0][1:]))

    # 获取修后恢复标准
    recover_Z_day = int(df_recover[df_recover['检修类型'] == 'Z']['修后恢复天数'].iloc[0])
    recover_Z_km = int(df_recover[df_recover['检修类型'] == 'Z']['修后恢复公里数'].iloc[0])
    recover_L_km = int(df_recover[df_recover['检修类型'] == 'L']['修后恢复公里数'].iloc[0])

    # 路线距离
    route_distance = dict(zip(df_route['交路'], df_route['distance']))

    # 初始剩余数据
    initial_Z_day = dict(zip(df_mileage['车组号'], df_mileage['Z剩余天数']))
    initial_Z_km = dict(zip(df_mileage['车组号'], df_mileage['Z剩余里程']))
    initial_L_km = dict(zip(df_mileage['车组号'], df_mileage['L剩余里程']))

    # 模型初始化
    model = cp_model.CpModel()

    # 决策变量定义
    x = {}  # x[v, r, d] ：车组v是否在d天执行交路r
    z = {}  # z[v, d]：车组v是否在d天进行Z检修
    l = {}  # l[v, d] ：车组v是否在d天进行L检修

    # 初始化所有Day1变量
    for v in vehicles:
        z[v, 'day1'] = model.NewConstant(0)
        l[v, 'day1'] = model.NewConstant(0)
        for r in routes:
            x[v, r, 'day1'] = model.NewConstant(0)

    # 根据df_day1的0/1矩阵设置Day1的固定值
    for _, row in df_day1.iterrows():
        route_or_maint = row['交路']  # 当前行是交路（r1-r35）还是检修（Z/L）
        for v in vehicles:
            if row[v] == 1:  # 车组v被分配到当前任务
                if route_or_maint == 'Z':
                    z[v, 'day1'] = model.NewConstant(1)
                elif route_or_maint == 'L':
                    l[v, 'day1'] = model.NewConstant(1)
                else:  # 是交路任务
                    x[v, route_or_maint, 'day1'] = model.NewConstant(1)

    for v in vehicles:
        for d in days[1:]:
            z[v, d] = model.NewBoolVar(f'z_{v}_{d}')
            l[v, d] = model.NewBoolVar(f'l_{v}_{d}')
            for r in routes:
                if r in vehicle_candidates[v]:
                    x[v, r, d] = model.NewBoolVar(f'x_{v}_{r}_{d}')
                else:
                    # 约束 6：候选交路限制
                    x[v, r, d] = model.NewConstant(0)

    # 约束 1：检修能力限制
    for d in days:
        model.Add(sum(z[v, d] for v in vehicles) <= capacity_Z[d])
        model.Add(sum(l[v, d] for v in vehicles) <= capacity_L[d])

    # 约束 2：每个交路每天必须有车组执行
    for r in routes:
        for d in days:
            if need_route[r, d] == 1:
                model.Add(sum(x[v, r, d] for v in vehicles) == 1)
            else:
                model.Add(sum(x[v, r, d] for v in vehicles) == 0)

    # 约束 3：每车组每天最多做一个任务
    for v in vehicles:
        for d in days:
            model.Add(sum(x[v, r, d] for r in routes) + z[v, d] + l[v, d] <= 1)

    # 约束 4：连续交路安排，每个车组v，每个R_ID下的连续交路对(r₁,r₂)，以及连续的时间对(t,t+1)要求 x[v,r₁,dayₜ] = x[v,r₂,dayₜ₊₁]
    for v in vehicles:
        for rid, r_list in r_id_routes.items():  # 遍历所有R_ID及其对应的交路列表
            if len(r_list) < 2:
                continue
            for i in range(len(r_list) - 1):  # 遍历交路列表中连续的每对交路
                r1, r2 = r_list[i], r_list[i + 1]
                for t in range(num_days - 1):
                    model.Add(x[v, r1, days[t]] == x[v, r2, days[t + 1]])

    # 约束 5：剩余天数/里程约束
    Z_day_left = {}
    Z_km_left = {}
    L_km_left = {}

    for v in vehicles:
        for t in range(num_days):
            d = days[t]

            # 初始化变量
            Z_day_left[v, d] = model.NewIntVar(0, 1000, f'Z_day_left_{v}_{d}')
            Z_km_left[v, d] = model.NewIntVar(0, 1000000, f'Z_km_left_{v}_{d}')
            L_km_left[v, d] = model.NewIntVar(0, 1000000, f'L_km_left_{v}_{d}')

            if t == 0:
                # Day1初始值（从数据读取）
                model.Add(Z_day_left[v, d] == initial_Z_day[v])
                model.Add(Z_km_left[v, d] == initial_Z_km[v])
                model.Add(L_km_left[v, d] == initial_L_km[v])
            else:
                d_prev = days[t - 1]

                # 创建布尔变量表示是否检修
                is_z = z[v, d]
                is_l = l[v, d]

                # Z剩余天数 = 若检修则重置为recover_Z_day，否则继续消耗
                model.Add(Z_day_left[v, d] == recover_Z_day).OnlyEnforceIf(is_z)
                model.Add(Z_day_left[v, d] == Z_day_left[v, d_prev] -
                          sum(x[v, r, d] for r in routes if r in vehicle_candidates[v])).OnlyEnforceIf(is_z.Not())

                # Z剩余里程 = 若检修则重置为recover_Z_km，否则继续消耗
                model.Add(Z_km_left[v, d] == recover_Z_km).OnlyEnforceIf(is_z)
                model.Add(Z_km_left[v, d] == Z_km_left[v, d_prev] -
                          sum(route_distance[r] * x[v, r, d] for r in routes if
                              r in vehicle_candidates[v])).OnlyEnforceIf(is_z.Not())

                # L剩余里程 = 若检修则重置为recover_L_km，否则继续消耗
                model.Add(L_km_left[v, d] == recover_L_km).OnlyEnforceIf(is_l)
                model.Add(L_km_left[v, d] == L_km_left[v, d_prev] -
                          sum(route_distance[r] * x[v, r, d] for r in routes if
                              r in vehicle_candidates[v])).OnlyEnforceIf(is_l.Not())

            # 非负约束
            model.Add(Z_day_left[v, d] >= 0)
            model.Add(Z_km_left[v, d] >= 0)
            model.Add(L_km_left[v, d] >= 0)

        # 目标函数
        #总检修次数
        total_z = model.NewIntVar(0, len(vehicles) * num_days, 'total_z')
        model.Add(total_z == sum(z[v, d] for v in vehicles for d in days))

        total_l = model.NewIntVar(0, len(vehicles) * num_days, 'total_l')
        model.Add(total_l == sum(l[v, d] for v in vehicles for d in days))
        #Z检修均衡
        z_count = [model.NewIntVar(0, len(vehicles), f'z_count_{d}') for d in days]
        for i, d in enumerate(days):
            model.Add(z_count[i] == sum(z[v, d] for v in vehicles))

        max_z = model.NewIntVar(0, len(vehicles), 'max_z')
        min_z = model.NewIntVar(0, len(vehicles), 'min_z')
        model.AddMaxEquality(max_z, z_count)
        model.AddMinEquality(min_z, z_count)
        z_variance = model.NewIntVar(0, len(vehicles), 'z_variance')
        model.Add(z_variance == max_z - min_z)

        #l检修均衡
        l_count = [model.NewIntVar(0, len(vehicles), f'l_count_{d}') for d in days]
        for i, d in enumerate(days):
            model.Add(l_count[i] == sum(l[v, d] for v in vehicles))

        max_l = model.NewIntVar(0, len(vehicles), 'max_l')
        min_l = model.NewIntVar(0, len(vehicles), 'min_l')
        model.AddMaxEquality(max_l, l_count)
        model.AddMinEquality(min_l, l_count)
        l_variance = model.NewIntVar(0, len(vehicles), 'l_variance')
        model.Add(l_variance == max_l - min_l)

    # 换车次数最小
    change_count = []
    for rid, r_list in r_id_routes.items():
        if len(r_list) < 2:  # 单交路不需要连续执行
            continue

        last_route = r_list[-1]  # 当前交路组的最后一个交路
        first_route = r_list[0]  # 当前交路组的第一个交路

        for t in range(1, num_days):  # 遍历连续天数对
            for v1 in vehicles:
                for v2 in vehicles:
                    if v1 != v2:
                        # 创建换车标记变量
                        change_var = model.NewBoolVar(f'change_{rid}_{days[t - 1]}_{days[t]}_{v1}_{v2}')

                        # 如果v1执行最后交路 且 v2执行首交路，则标记为换车
                        model.AddBoolAnd([
                            x[v1, last_route, days[t - 1]],
                            x[v2, first_route, days[t]]
                        ]).OnlyEnforceIf(change_var)

                        # 否则不标记为换车
                        model.AddBoolOr([
                            x[v1, last_route, days[t - 1]].Not(),
                            x[v2, first_route, days[t]].Not()
                        ]).OnlyEnforceIf(change_var.Not())

                        change_count.append(change_var)

    # obj = model.NewIntVar(0, 1000000, 'obj')
    # weight_total_maint = 10  # 总检修次数权重
    # weight_balance = 1 # 检修均衡性权重
    # weight_change = 5  # 换车次数权重
    # model.Add(
    #     obj ==
    #     weight_total_maint * (total_z + total_l) +  # 最小化总检修次数
    #     weight_balance * (z_variance + l_variance ) +  # 最小化检修不均衡
    #     weight_change * sum(change_count)  # 最小化换车次数
    # )
    # model.Minimize(obj)
    model.Minimize(sum(change_count))
    solver = cp_model.CpSolver()
    solver.parameters.max_time_in_seconds = 600
    status = solver.Solve(model)

    # 打印结果
    if status == cp_model.OPTIMAL or status == cp_model.FEASIBLE:
        print("可行解找到")
        # print(f"目标函数值: {solver.Value(obj)}")
        # print(f"总检修次数: {solver.Value(total_z)} + {solver.Value(total_l)}")
        # print(f"Z检修不均衡: {solver.Value(z_variance)}")
        # print(f"L检修不均衡: {solver.Value(l_variance)}")
        print(f"换车次数: {sum(solver.Value(v) for v in change_count)}")

        export_to_excel(
            solver=solver,
            x=x,
            z=z,
            l=l,
            vehicles=vehicles,
            routes=routes,
            days=days,
            filename="railway_schedule_result.xlsx"
        )
        # for d in days:
        #     print(f"\n📅 {d} 日排班:")
        #     for r in routes:
        #         for v in vehicles:
        #             if solver.Value(x[v, r, d]):
        #                 print(f"  - 交路 {r} 由车组 {v} 执行")
        #     for v in vehicles:
        #         if solver.Value(z[v, d]):
        #             print(f"  - 车组 {v} 进行 Z 检修")
        #         if solver.Value(l[v, d]):
        #             print(f"  - 车组 {v} 进行 L 检修")
    else:
        print("未找到可行解")


if __name__ == "__main__":
    # main()
    data_file = "railway_schedule_result.xlsx"
    calc_excel(data_file )
    calc_excel("Result.xlsx")


